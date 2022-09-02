from openpyxl import workbook
from openpyxl import load_workbook

xlBook = load_workbook(filename="D:/Classes/Suraj/ShaftData.xlsx")
#print(xlBook.sheetnames)

xlSheet = xlBook['Shaft Parameters']
# print(xlSheet.title)

# xlSheet = xlBook.active
# print(xlSheet.title)

# print(xlSheet['C6'].value)
# print(xlSheet['D6'].value)
#
# print(xlSheet['C7'].value)
# print(xlSheet['D7'].value)

# print(xlSheet.cell(row=6, column=3).value)
# print(xlSheet.cell(row=6, column=4).value)

# print(xlSheet['C6:D7'])
# print(xlSheet['C6:D7'][0])
# print(xlSheet['C6:D7'][1])
#
# print(xlSheet['C6:D7'][0][0])
# print(xlSheet['C6:D7'][0][0].value)

xlSheet['C5'] = 'Hello'
print(xlSheet['C5'].value)
